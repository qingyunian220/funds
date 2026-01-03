import re

import pandas as pd
import akshare as ak
import time

import requests
from bs4 import BeautifulSoup
from tqdm import tqdm
import schedule
import threading
from flask import Flask, jsonify, request, render_template
import os
import json
import numpy as np
from openpyxl import load_workbook

from fund_data_processor import get_fund_name_by_code
from fund_search_parser import fetch_and_parse_fund_search
from jiuquan_fund import parse_fund_data

# # Flask应用配置
# app = Flask(__name__)
# app.config['JSON_AS_ASCII'] = False
#
# # 设置Excel文件路径
# EXCEL_FILE = 'index-fund.xlsx'

def get_fund_info(fund_code):
    """
    通过天天基金接口获取基金的成立时间和最新规模（优化版）

    该方法只请求一次基金详情页，并使用 BeautifulSoup 解析 HTML，
    同时提取成立日期和最新规模，效率更高。

    参数
    ----------
    fund_code : str
        基金代码，如 '001186'

    返回
    -------
    dict
        {
            '基金代码': str,
            '成立日期': str,
            '最新规模': str,
            '规模日期': str,
            '错误': str  # 如果获取失败
        }
    """
    result = {
        '基金代码': fund_code,
        '成立日期': '',
        '最新规模': '',
        '规模日期': '',
        '错误': None
    }

    # 请求基金详情页
    detail_url = f'http://fund.eastmoney.com/{fund_code}.html'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
    }

    try:
        response = requests.get(detail_url, headers=headers, timeout=10)
        response.encoding = 'utf-8'

        # 如果请求失败（如404），则直接返回
        if response.status_code != 200:
            result['错误'] = f"请求失败，状态码: {response.status_code}"
            return result

        # 使用 BeautifulSoup 解析 HTML
        soup = BeautifulSoup(response.text, 'html.parser')

        # 查找包含基金信息的表格
        info_table = soup.find('div', class_='infoOfFund')
        if not info_table:
            result['错误'] = "未找到基金信息区域（.infoOfFund）"
            return result

        # 遍历表格中的所有单元格
        for td in info_table.find_all('td'):
            td_text = td.get_text(strip=True)

            # 提取成立日期
            if '成 立 日' in td_text:
                # 使用正则表达式提取日期，更精确
                date_match = re.search(r'(\d{4}-\d{2}-\d{2})', td_text)
                if date_match:
                    result['成立时间'] = date_match.group(1)
            # 提取规模和规模日期
            if '规模' in td_text and '亿元' in td_text:
                # 使用正则表达式提取规模和日期
                scale_match = re.search(r'([\d.]+)亿元（(\d{4}-\d{2}-\d{2})）', td_text)
                if scale_match:
                    result['最新规模'] = f"{scale_match.group(1)}亿元"
                    result['最新规模日期'] = scale_match.group(2)

    except requests.exceptions.RequestException as e:
        result['错误'] = f"网络请求错误: {e}"
    except Exception as e:
        result['错误'] = f"解析错误: {e}"

    return result

def fetch_fund_data(fund_type):
    """获取指定类型基金数据（只保留各时间段前10名的基金）"""
    # 获取所有基金基础信息
    fund_open_fund_rank_em_df = ak.fund_open_fund_rank_em(symbol="全部")
    fund_df = fund_open_fund_rank_em_df[fund_open_fund_rank_em_df["基金简称"].str.contains(fund_type, na=False)]
    fund_df = fund_df[fund_df["基金简称"].str.contains("C", na=False)].copy()

    # 过滤掉"近6月"为空的数据
    fund_df = fund_df.dropna(subset=['近6月']).copy()

    exclude_keywords = ["红利", "基本面", "价值", "非银", "成长", "低波动","信息技术","周期","非周期","地产","有色","医药","保险","金融","持有"]
    for keyword in tqdm(exclude_keywords, desc="过滤关键词"):
        fund_df = fund_df[~fund_df["基金简称"].str.contains(keyword, na=False)].copy()

    # 筛选出在各时间段任意一个进入前10的基金
    return_columns = ['近1月', '近3月', '近6月', '近1年', '今年来']
    top_indices = set()
    for col in return_columns:
        if col in fund_df.columns:
            top_10_idx = fund_df[col].nlargest(10).index
            top_indices.update(top_10_idx)

    fund_df = fund_df.loc[list(top_indices)].copy()
    print(f"{fund_type}基金: 从全部基金中筛选出 {len(fund_df)} 只各时间段前10基金")

    fund_df.loc[:, "成立时间"] = ""
    fund_df.loc[:, "最新规模"] = ""
    fund_df.loc[:, "换手率"] = ""
    fund_df.loc[:, "前10大重仓股占比"] = ""
    fund_df.loc[:, "持股行业集中度"] = ""

    for idx, row in tqdm(fund_df.iterrows(), total=fund_df.shape[0], desc=f"获取{fund_type}基金详情"):
        code = row["基金代码"]
        # 1. 根据基金代码查找基金名称
        fund_name = get_fund_name_by_code(str(code))
        if not fund_name:
            print(f"无法获取基金 {code} 的名称")
            return None
        base_name = fund_name
        if fund_name.endswith('A') or fund_name.endswith('C'):
            base_name = fund_name[:-1]
        # 2. 根据基金名称查找对应的基金代码(A+C)
        # [{'code': '015381', 'name': '东方兴瑞趋势领航混合A'}, {'code': '015382', 'name': '东方兴瑞趋势领航混合C'}]
        code_names = fetch_and_parse_fund_search(base_name)
        # 3. 获取A类和C类基金的规模数据
        for code_name in code_names:
            info = get_fund_info(code_name['code'])
            # 解析并累加规模数据
            scale_match = re.search(r'([\d.]+)亿元', info['最新规模'])
            if scale_match:
                scale_value = float(scale_match.group(1))
                # 如果当前值是空字符串，则初始化为0
                current_scale = fund_df.loc[idx, "最新规模"]
                if current_scale is not None == "":
                    current_scale = 0
                else:
                    # 提取当前值中的数字部分
                    current_match = re.search(r'([\d.]+)亿元', str(current_scale))
                    if current_match:
                        current_scale = float(current_match.group(1))
                    else:
                        current_scale = 0
                # 累加规模值
                new_scale = current_scale + scale_value
                fund_df.loc[idx, "最新规模"] = f"{new_scale:.2f}亿元"
        try:
            info = get_fund_info(code)
            fund_df.loc[idx, "成立时间"] = info['成立时间']
            # fund_df.loc[idx, "最新规模"] = info['最新规模']
            # 获取换手率和重仓股信息
            fund_detail = parse_fund_data(code)
            if fund_detail:
                if '换手率' in fund_detail:
                    fund_df.loc[idx, "换手率"] = fund_detail['换手率']
                if '前10大重仓股占比' in fund_detail:
                    fund_df.loc[idx, "前10大重仓股占比"] = fund_detail['前10大重仓股占比']
                if '持股行业集中度' in fund_detail:
                    fund_df.loc[idx, "持股行业集中度"] = fund_detail['持股行业集中度']
        except Exception as e:
            print(f"基金代码{code}查询失败: {e}")

    return fund_df.sort_values(by='近6月', ascending=False)

def fetch_small_fund_data():
    """获取小微盘基金数据（只保留各时间段前10名的基金）"""
    # 从Excel文件中读取基金代码，确保基金代码作为字符串处理
    small_funds_df = pd.read_excel('small_funds.xlsx', dtype={'code': str})

    # 获取所有基金基础信息
    fund_open_fund_rank_em_df = ak.fund_open_fund_rank_em(symbol="全部")

    # 筛选出在small_funds.xlsx中的基金
    fund_df = fund_open_fund_rank_em_df[fund_open_fund_rank_em_df["基金代码"].isin(small_funds_df["code"])].copy()

    # 筛选出在各时间段任意一个进入前10的基金
    return_columns = ['近1月', '近3月', '近6月', '近1年', '今年来']
    top_indices = set()
    for col in return_columns:
        if col in fund_df.columns:
            top_10_idx = fund_df[col].nlargest(10).index
            top_indices.update(top_10_idx)

    fund_df = fund_df.loc[list(top_indices)].copy()
    print(f"小微盘基金: 从全部基金中筛选出 {len(fund_df)} 只各时间段前10基金")

    fund_df.loc[:, "成立时间"] = ""
    fund_df.loc[:, "最新规模"] = ""
    fund_df.loc[:, "换手率"] = ""
    fund_df.loc[:, "前10大重仓股占比"] = ""
    fund_df.loc[:, "持股行业集中度"] = ""

    for idx, row in tqdm(fund_df.iterrows(), total=fund_df.shape[0], desc="获取小微盘基金详情"):
        code = row["基金代码"]
        # 1. 根据基金代码查找基金名称
        fund_name = get_fund_name_by_code(str(code))
        if not fund_name:
            print(f"无法获取基金 {code} 的名称")
            return None
        base_name = fund_name
        if fund_name.endswith('A') or fund_name.endswith('C'):
            base_name = fund_name[:-1]
        # 2. 根据基金名称查找对应的基金代码(A+C)
        # [{'code': '015381', 'name': '东方兴瑞趋势领航混合A'}, {'code': '015382', 'name': '东方兴瑞趋势领航混合C'}]
        code_names = fetch_and_parse_fund_search(base_name)
        # 3. 获取A类和C类基金的规模数据
        for code_name in code_names:
            info = get_fund_info(code_name['code'])

            # 解析并累加规模数据
            scale_match = re.search(r'([\d.]+)亿元', info['最新规模'])
            if scale_match:
                scale_value = float(scale_match.group(1))
                # 如果当前值是空字符串，则初始化为0
                current_scale = fund_df.loc[idx, "最新规模"]
                if current_scale is not None == "":
                    current_scale = 0
                else:
                    # 提取当前值中的数字部分
                    current_match = re.search(r'([\d.]+)亿元', str(current_scale))
                    if current_match:
                        current_scale = float(current_match.group(1))
                    else:
                        current_scale = 0
                # 累加规模值
                new_scale = current_scale + scale_value
                fund_df.loc[idx, "最新规模"] = f"{new_scale:.2f}亿元"
        try:
            info = get_fund_info(code)
            fund_df.loc[idx, "成立时间"] = info['成立时间']
            # fund_df.loc[idx, "最新规模"] = info['最新规模']
            # 获取换手率和重仓股信息
            fund_detail = parse_fund_data(code)
            if fund_detail:
                if '换手率' in fund_detail:
                    fund_df.loc[idx, "换手率"] = fund_detail['换手率']
                if '前10大重仓股占比' in fund_detail:
                    fund_df.loc[idx, "前10大重仓股占比"] = fund_detail['前10大重仓股占比']
                if '持股行业集中度' in fund_detail:
                    fund_df.loc[idx, "持股行业集中度"] = fund_detail['持股行业集中度']
        except Exception as e:
            print(f"基金代码{code}查询失败: {e}")

    return fund_df.sort_values(by='近6月', ascending=False)

def highlight_top_50_all_columns(df):
    """为表格数据添加样式标记"""
    # 创建一个样式DataFrame，默认为空字符串（无样式）
    styles = pd.DataFrame('', index=df.index, columns=df.columns)

    # 定义收益率列（去掉近1周）
    return_columns = ['近1月', '近3月', '近6月', '近1年', '今年来']

    # 记录每个基金在多少个收益率列中进入前10
    top_count = {idx: 0 for idx in df.index}

    # 对每个收益率列，标记其前10名
    for col in return_columns:
        if col in df.columns:
            # 获取该列排序后的前10个索引
            top_10_idx = df[col].nlargest(10).index
            # 将对应位置设置为黄色背景
            styles.loc[top_10_idx, col] = 'background-color: yellow'
            # 更新每个基金进入前10的次数
            for idx in top_10_idx:
                top_count[idx] += 1

    # 对至少有3列进入前10的基金，将基金简称设置为金黄色背景
    for idx, count in top_count.items():
        if count >= 3:  # 至少有3列进入前10（共5个时间段）
            # 标注基金简称为金黄色
            styles.loc[idx, '基金简称'] = 'background-color: gold'

    return styles

def save_to_excel(writer, fund_df, sheet_name):
    """保存数据到Excel文件"""
    if not fund_df.empty:  # 确保DataFrame不为空
        styled_df = fund_df.style.apply(highlight_top_50_all_columns, axis=None)
        styled_df.to_excel(writer, sheet_name=sheet_name, index=False)

def highlight_excess_returns(df):
    """为超额收益率表格添加样式标记"""
    # 创建一个样式DataFrame，默认为空字符串（无样式）
    styles = pd.DataFrame('', index=df.index, columns=df.columns)

    # 定义超额收益率列
    excess_return_columns = ['近1周超额', '近1月超额', '近3月超额', '近6月超额', '近1年超额', '近2年超额', '近3年超额', '今年来超额']

    # 记录每个基金在多少个超额收益率列中进入前10
    top_count = {idx: 0 for idx in df.index}

    # 对每个超额收益率列，标记其前10名
    for col in excess_return_columns:
        if col in df.columns:
            # 获取该列排序后的前10个索引
            top_10_idx = df[col].nlargest(10).index
            # 将对应位置设置为黄色背景
            styles.loc[top_10_idx, col] = 'background-color: yellow'
            # 更新每个基金进入前10的次数
            for idx in top_10_idx:
                top_count[idx] += 1

    # 对至少有3列进入前10的基金，将基金简称设置为金黄色背景
    for idx, count in top_count.items():
        if count >= 3:  # 至少有3列进入前10
            # 标注基金简称为金黄色
            styles.loc[idx, '基金简称'] = 'background-color: gold'

    return styles

def calculate_excess_returns(writer):
    """计算超额收益率并保存到Excel文件"""
    # 定义基金类型与基准基金代码的映射关系
    benchmark_map = {
        "沪深300": "510300",
        "中证500": "512510",
        "A500":"563360",
        "中证800": "515810",
        "中证1000": "516300",
        "中证2000": "563300",
        "国证2000": "159907"
    }
    # 定义收益率列
    return_columns = ['近1周', '近1月', '近3月', '近6月', '近1年', '近2年', '近3年', '今年来']
    # 获取基金排名数据
    fund_exchange_rank_em_df = ak.fund_exchange_rank_em()
    fund_open_fund_rank_em_df = ak.fund_open_fund_rank_em(symbol="混合型")
    # 为每种基金类型计算超额收益率
    for fund_type, benchmark_code in benchmark_map.items():
        print(f"正在处理{fund_type}基金的超额收益率，基准基金代码：{benchmark_code}")
        
        # 读取已保存的基金数据
        try:
            fund_df = pd.read_excel('index-fund.xlsx', sheet_name=f'{fund_type}基金')
        except:
            print(f"无法读取{fund_type}基金数据")
            continue
        
        # 获取基准基金数据
        benchmark_df = fund_exchange_rank_em_df[fund_exchange_rank_em_df["基金代码"] == benchmark_code]
        
        if benchmark_df.empty:
            print(f"未找到{fund_type}基金的基准基金{benchmark_code}")
            continue
            
        # 获取基准收益率
        benchmark_returns = {}
        for col in return_columns:
            if col in benchmark_df.columns:
                # 处理数值数据，确保正确转换
                value = benchmark_df[col].iloc[0]
                if pd.isna(value):
                    benchmark_returns[col] = 0
                else:
                    # 如果是字符串，去除%符号并转换为数值
                    if isinstance(value, str):
                        benchmark_returns[col] = pd.to_numeric(value.rstrip('%'), errors='coerce') / 100
                    else:
                        # 如果已经是数值类型，直接使用
                        benchmark_returns[col] = pd.to_numeric(value, errors='coerce') / 100
            else:
                benchmark_returns[col] = 0
        
        # 计算超额收益率
        for col in return_columns:
            if col in fund_df.columns:
                # 创建超额收益率列名
                excess_col = f'{col}超额'
                # 将原始收益率转换为数值
                fund_df[col] = pd.to_numeric(fund_df[col].astype(str).str.rstrip('%'), errors='coerce') / 100
                # 计算超额收益率（保持数值格式）
                fund_df[excess_col] = (fund_df[col] - benchmark_returns[col])*100
        
        # 只保留指定的列
        columns_to_keep = ["基金代码", "基金简称", "日期", "近1周超额", "近1月超额", "近3月超额", "近6月超额", 
                          "近1年超额", "近2年超额", "近3年超额", "今年来超额", "成立时间", "最新规模", "换手率",
                           "前10大重仓股占比", "持股行业集中度"]
        # 检查哪些列实际存在于DataFrame中
        existing_columns = [col for col in columns_to_keep if col in fund_df.columns]
        fund_df = fund_df[existing_columns]
        
        # 确保基金代码为6位，不足的向前填充0
        if "基金代码" in fund_df.columns:
            fund_df["基金代码"] = fund_df["基金代码"].astype(str).str.zfill(6)
        
        # 保存带有超额收益率的数据，并应用样式
        styled_df = fund_df.style.apply(highlight_excess_returns, axis=None)
        styled_df.to_excel(writer, sheet_name=f'{fund_type}基金_超额', index=False)
    
    # 单独处理小微盘基金
    try:
        small_fund_df = pd.read_excel('index-fund.xlsx', sheet_name='小微盘')
        # 小微盘基金使用320016作为基准
        benchmark_code = "320016"
        benchmark_df = fund_open_fund_rank_em_df[fund_open_fund_rank_em_df["基金代码"] == benchmark_code]
        
        if not benchmark_df.empty:
            # 获取基准收益率
            benchmark_returns = {}
            for col in return_columns:
                if col in benchmark_df.columns:
                    # 处理数值数据，确保正确转换
                    value = benchmark_df[col].iloc[0]
                    if pd.isna(value):
                        benchmark_returns[col] = 0
                    else:
                        # 如果是字符串，去除%符号并转换为数值
                        if isinstance(value, str):
                            benchmark_returns[col] = pd.to_numeric(value.rstrip('%'), errors='coerce') / 100
                        else:
                            # 如果已经是数值类型，直接使用
                            benchmark_returns[col] = pd.to_numeric(value, errors='coerce') / 100
                else:
                    benchmark_returns[col] = 0
            
            # print(f"小微盘基准基金{benchmark_code}收益率: {benchmark_returns}")
            
            # 计算超额收益率
            for col in return_columns:
                if col in small_fund_df.columns:
                    # 创建超额收益率列名
                    excess_col = f'{col}超额'
                    # 将原始收益率转换为数值
                    small_fund_df[col] = pd.to_numeric(small_fund_df[col].astype(str).str.rstrip('%'), errors='coerce') / 100
                    # 计算超额收益率（保持数值格式）
                    small_fund_df[excess_col] = (small_fund_df[col] - benchmark_returns[col])*100
            
            # 只保留指定的列
            columns_to_keep = ["基金代码", "基金简称", "日期", "近1周超额", "近1月超额", "近3月超额", "近6月超额", 
                              "近1年超额", "近2年超额", "近3年超额", "今年来超额", "成立时间", "最新规模", "换手率",
                               "前10大重仓股占比", "持股行业集中度"]
            # 检查哪些列实际存在于DataFrame中
            existing_columns = [col for col in columns_to_keep if col in small_fund_df.columns]
            small_fund_df = small_fund_df[existing_columns]
            
            # 确保基金代码为6位，不足的向前填充0
            if "基金代码" in small_fund_df.columns:
                small_fund_df["基金代码"] = small_fund_df["基金代码"].astype(str).str.zfill(6)
            
            # 保存带有超额收益率的数据，并应用样式
            styled_df = small_fund_df.style.apply(highlight_excess_returns, axis=None)
            styled_df.to_excel(writer, sheet_name='小微盘_超额', index=False)
        else:
            print(f"未找到小微盘基金的基准基金{benchmark_code}")
    except Exception as e:
        print(f"处理小微盘基金超额收益率时出错: {e}")
        import traceback
        traceback.print_exc()

def adjust_column_width(filename):
    """调整Excel文件的列宽"""
    # 打开已存在的Excel文件
    workbook = load_workbook(filename)
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # 限制最大宽度为50
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    workbook.save(filename)

def update_fund_data():
    """更新基金数据的函数"""
    print(f"开始更新基金数据: {time.strftime('%Y-%m-%d %H:%M:%S')}")
    
    try:
        # 创建一个ExcelWriter对象
        filename = f'index-fund.xlsx'
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            fund_types = ["沪深300", "中证500","A500","中证800", "中证1000", "中证2000","国证2000"]
            for fund_type in fund_types:
                fund_df = fetch_fund_data(fund_type)
                save_to_excel(writer, fund_df, f'{fund_type}基金')
            
            # 添加小微盘基金数据
            small_fund_df = fetch_small_fund_data()
            save_to_excel(writer, small_fund_df, '小微盘')
        
        # 计算超额收益率并保存到新的工作表
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
            calculate_excess_returns(writer)
        
        # 调整列宽
        try:
            adjust_column_width(filename)
        except Exception as e:
            print(f"调整列宽时出错: {e}")
        
        print(f"已将所有C份额基金的排序结果保存为'{filename}'，每个时间段的前10名标黄，至少有4个时间段进入前10的基金其简称标金黄色。")
        print(f"基金数据更新完成: {time.strftime('%Y-%m-%d %H:%M:%S')}")
    except Exception as e:
        print(f"更新基金数据时发生错误: {e}")
        import traceback
        traceback.print_exc()

def run_scheduler():
    """运行定时任务"""
    while True:
        try:
            schedule.run_pending()
        except Exception as e:
            print(f"调度器运行时发生错误: {e}")
        time.sleep(60)  # 每分钟检查一次


def start_scheduler():
    """启动定时任务"""
    # 立即执行一次更新
    update_fund_data()
    
    # 设置每天凌晨12:00执行更新
    schedule.every().day.at("00:00").do(update_fund_data)
    
    print("定时任务已启动，每天凌晨12:00将自动更新数据。")
    print("按 Ctrl+C 可以停止程序。")
    
    # 在单独的线程中运行定时任务
    scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
    scheduler_thread.start()
    
    # 保持主线程运行
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("程序已被用户中断")
    except Exception as e:
        print(f"主线程发生错误: {e}")

if __name__ == '__main__':
    # 启动定时任务
    start_scheduler()