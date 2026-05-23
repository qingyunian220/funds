import re
import numpy as np
from concurrent.futures import ThreadPoolExecutor, as_completed

import json
import re
import requests
from typing import Dict, List, Optional, Any
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import gzip
import zlib
import brotli
import pandas as pd
import akshare as ak
from datetime import datetime, timedelta
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from bs4 import BeautifulSoup

HEADER_JIUQUAN = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
    'Content-Type': 'application/json',
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
    'Referer': 'https://apiv2.jiucaishuo.com/'
}


def get_industry_concentration_from_akshare(fund_code):
    """
    通过AKShare获取基金的行业配置数据，并获取第一大持仓行业占比

    参数
    ------
    fund_code : str
        基金代码

    返回
    ------
    str or None
        第一大持仓行业占比（百分比字符串），获取失败返回None
    """
    try:
        # 调用AKShare的行业配置接口
        industry_df = ak.fund_portfolio_industry_allocation_em(symbol=fund_code)

        if industry_df.empty:
            return None

        # 获取最新一期的数据（第一行）
        latest_data = industry_df.iloc[0]

        # 获取所有行业占比列（排除报告期列）
        industry_cols = [col for col in industry_df.columns if col not in ['报告期']]

        # 提取行业占比
        industry_ratios = []
        for col in industry_cols:
            ratio = latest_data[col]
            if pd.notna(ratio) and ratio > 0:
                industry_ratios.append(ratio)

        if not industry_ratios:
            return None

        # 按降序排序
        industry_ratios.sort(reverse=True)

        # 取第一大行业的占比
        top1_ratio = industry_ratios[0]

        # 返回百分比格式，例如：'第一大持仓行业占比87.36%'
        return f"第一大持仓行业占比{top1_ratio:.2f}%"

    except Exception as e:
        # print(f"AKShare获取基金 {fund_code} 行业集中度失败: {e}")
        return None

def create_session():
    """
    创建一个带有重试策略的会话
    """
    session = requests.Session()

    # 定义重试策略
    retry_strategy = Retry(
        total=3,  # 总重试次数
        backoff_factor=1,  # 重试间隔
        status_forcelist=[429, 500, 502, 503, 504],  # 需要重试的状态码
    )

    # 创建适配器并挂载到会话
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("http://", adapter)
    session.mount("https://", adapter)

    return session

def decompress_response_content(response):
    """
    尝试解压响应内容
    """
    content = response.content
    headers = response.headers

    # 检查Content-Encoding
    content_encoding = headers.get('Content-Encoding', '').lower()

    try:
        if 'gzip' in content_encoding:
            content = gzip.decompress(content)
        elif 'deflate' in content_encoding:
            content = zlib.decompress(content)
        elif 'br' in content_encoding:
            content = brotli.decompress(content)
    except Exception as e:
        print(f"解压响应内容失败: {str(e)}")
        return response.content
    return content

def extract_numeric_value(text):
    """
    从文本中提取数字和单位，例如从"基金最新一期规模5692.07万"提取"5692.07万"
    """
    if not isinstance(text, str):
        return text

    # 使用正则表达式匹配数字和单位（包括小数点、百分比符号、中文单位等）
    pattern = r'(\d+(?:\.\d+)?(?:[%万亿亿千万百十])*)'
    matches = re.findall(pattern, text)

    if matches:
        return matches[0]
    return text


def parse_fund_data(fund_code):
    """
    调用API接口并解析基金数据
    """
    url = "https://apiv2.jiucaishuo.com/funddetail/detail/fund-high-lights"
    payload = {
        "fund_code": fund_code,
        "type": "h5"
    }

    # 创建会话
    session = create_session()

    try:
        response = session.post(url, json=payload, headers=HEADER_JIUQUAN, timeout=15, stream=True)
        response.raise_for_status()

        # 尝试解压响应内容
        # content = decompress_response_content(response)
        content = response.content

        # 尝试不同的编码方式解码
        try:
            text = content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                text = content.decode('gbk')
            except UnicodeDecodeError:
                text = content.decode('utf-8', errors='ignore')

        # 清理文本内容
        text = text.strip()

        # 检查响应内容是否为空
        if not text:
            print(f"基金 {fund_code} 接口返回空内容")
            return None

        # 检查响应内容是否为有效的JSON
        try:
            # 尝试解析JSON之前，先检查内容是否看起来像JSON
            if not (text.startswith('{') or text.startswith('[')):
                print(f"基金 {fund_code} 接口返回非JSON内容: {text[:100]}...")
                return None

            data = json.loads(text)
        except json.JSONDecodeError as e:
            print(f"基金 {fund_code} JSON解析失败: {str(e)}")
            print(f"响应状态码: {response.status_code}")
            print(f"响应头: {dict(response.headers)}")
            print(f"响应内容前200字符: {text[:200]!r}")
            # 将响应内容保存到文件供分析
            with open(f'{fund_code}_response.txt', 'w', encoding='utf-8') as f:
                f.write(text)
            print(f"响应内容已保存到 {fund_code}_response.txt")
            return None

        # print("接口返回数据:", data)

        if data['code'] != 0:
            print(f"基金 {fund_code} 接口返回错误: {data['message']}")
            return None

        return parse_fund_details(data['data'], fund_code)

    except requests.exceptions.RequestException as e:
        print(f"基金 {fund_code} 请求失败: {str(e)}")
        return None
    except Exception as e:
        print(f"基金 {fund_code} 解析异常: {str(e)}")
        return None


def parse_fund_details(data, fund_code):
    """
    解析基金详细信息
    """
    result = {
        '基金代码': fund_code,
        '解析时间': pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
    }

    # 解析持仓特征
    position_features = data.get('tssj_list', [])
    for feature in position_features:
        if feature.get('name') == '持仓特征':
            for tag in feature.get('tags', []):
                left_title = tag.get('left_title', '')
                info = tag.get('info', '')
                if '换手率' in left_title:
                    result['换手率'] = extract_numeric_value(info)
                elif '持股集中度' in left_title:
                    result['前10大重仓股占比'] = extract_numeric_value(info)
                elif '持股行业集中度' in left_title:
                    result['持股行业集中度'] = info  # 保持原样

    # 如果久财说没有返回持股行业集中度，尝试从AKShare获取
    if '持股行业集中度' not in result or not result['持股行业集中度']:
        akshare_result = get_industry_concentration_from_akshare(fund_code)
        if akshare_result:
            result['持股行业集中度'] = akshare_result

    # 解析基金经理信息
    manager_features = data.get('tssj_list', [])
    for feature in manager_features:
        if feature.get('name') == '基金经理':
            for tag in feature.get('tags', []):
                left_title = tag.get('left_title', '')
                info = tag.get('info', '')
                if '管理总规模' in left_title:
                    result['管理总规模'] = info


    return result

def fetch_and_parse_fund_search(keyword: str) -> List[Dict[str, str]]:
    """
    根据关键词获取并解析基金搜索数据

    Args:
        keyword (str): 搜索关键词

    Returns:
        list: 解析后的基金数据列表
    """
    url = "https://fundsuggest.eastmoney.com/FundSearch/api/FundSearchAPI.ashx"

    params = {
        "m": "1",
        "key": keyword
    }

    try:
        response = requests.get(url, params=params)
        response.raise_for_status()
        return parse_fund_search_response(response.text)
    except Exception as e:
        return {
            "error": True,
            "error_code": -1,
            "error_message": f"请求失败: {str(e)}"
        }


def parse_fund_search_response(response_text: str) -> list:
    """
    解析基金搜索API的响应数据

    Args:
        response_text (str): API返回的原始响应文本

    Returns:
        dict: 解析后的基金数据
    """
    try:
        # 首先尝试解析为JSON（直接JSON格式）
        data = json.loads(response_text)
    except json.JSONDecodeError:
        # 如果失败，尝试去除JSONP包装后再解析
        json_str = re.search(r'^[^(]*\((.*)\)$', response_text.strip())
        if not json_str:
            raise ValueError("无效的响应格式")
        data = json.loads(json_str.group(1))

    # 如果有错误码且不为0，返回错误信息
    if data.get("ErrCode", 0) != 0:
        return {
            "error": True,
            "error_code": data.get("ErrCode"),
            "error_message": data.get("ErrMsg", "未知错误")
        }

    # 解析基金数据
    parsed_funds = []
    for fund_data in data.get("Datas", []):
        parsed_fund = {
            # 基金基本信息
            "code": fund_data.get("CODE"),
            "name": fund_data.get("NAME"),
        }

        parsed_funds.append(parsed_fund)
    return parsed_funds

def get_fund_info(fund_code):
    """
    通过天天基金接口获取基金的成立时间和最新规模（优化版）

    该方法只请求一次基金详情页，并使用 BeautifulSoup 解析 HTML，
    同时提取成立日期和最新规模，效率更高。

    参数
    ----------
    fund_code : str
        基金代码，如 '001186'

    返回
    -------
    dict
        {
            '基金代码': str,
            '成立日期': str,
            '最新规模': str,
            '规模日期': str,
            '错误': str  # 如果获取失败
        }
    """
    result = {
        '基金代码': fund_code,
        '成立日期': '',
        '最新规模': '',
        '规模日期': '',
        '错误': None
    }

    # 请求基金详情页
    detail_url = f'http://fund.eastmoney.com/{fund_code}.html'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
    }

    try:
        response = requests.get(detail_url, headers=headers, timeout=10)
        response.encoding = 'utf-8'

        # 如果请求失败（如404），则直接返回
        if response.status_code != 200:
            result['错误'] = f"请求失败，状态码: {response.status_code}"
            return result

        # 使用 BeautifulSoup 解析 HTML
        soup = BeautifulSoup(response.text, 'html.parser')

        # 查找包含基金信息的表格
        info_table = soup.find('div', class_='infoOfFund')
        if not info_table:
            result['错误'] = "未找到基金信息区域（.infoOfFund）"
            return result

        # 遍历表格中的所有单元格
        for td in info_table.find_all('td'):
            td_text = td.get_text(strip=True)

            # 提取成立日期
            if '成 立 日' in td_text:
                # 使用正则表达式提取日期，更精确
                date_match = re.search(r'(\d{4}-\d{2}-\d{2})', td_text)
                if date_match:
                    result['成立时间'] = date_match.group(1)
            # 提取规模和规模日期
            if '规模' in td_text and '亿元' in td_text:
                # 使用正则表达式提取规模和日期
                scale_match = re.search(r'([\d.]+)亿元（(\d{4}-\d{2}-\d{2})）', td_text)
                if scale_match:
                    result['最新规模'] = f"{scale_match.group(1)}亿元"
                    result['最新规模日期'] = scale_match.group(2)

    except requests.exceptions.RequestException as e:
        result['错误'] = f"网络请求错误: {e}"
    except Exception as e:
        result['错误'] = f"解析错误: {e}"

    return result

def get_top10_stocks_weight_robust(fund_code):
    """
    获取指定基金最近季度前十大重仓股的占净值比例之和
    使用多种方法处理编码问题
    """
    current_year = str(datetime.now().year)
    previous_year = str(datetime.now().year - 1)

    # 尝试不同方法获取数据
    fund_portfolio_hold_em_df = None

    # 方法1: 直接使用akshare，处理可能的编码错误
    try:
        fund_portfolio_hold_em_df = ak.fund_portfolio_hold_em(symbol=fund_code, date=current_year)
    except KeyError as e:
        try:
            fund_portfolio_hold_em_df = ak.fund_portfolio_hold_em(symbol=fund_code, date=previous_year)
        except KeyError as e2:
            return None

    if fund_portfolio_hold_em_df is None or fund_portfolio_hold_em_df.empty:
        return None

    # 获取所有不同的季度
    try:
        quarters = fund_portfolio_hold_em_df['季度'].unique()
    except KeyError:
        possible_date_cols = [col for col in fund_portfolio_hold_em_df.columns if '季度' in col or 'date' in col.lower() or 'period' in col.lower()]
        if possible_date_cols:
            quarters = fund_portfolio_hold_em_df[possible_date_cols[0]].unique()
        else:
            return None

    # 确定最新的季度
    latest_quarter = sorted(quarters, reverse=True)[0]

    # 提取最近一个季度的数据
    try:
        filtered_df = fund_portfolio_hold_em_df[fund_portfolio_hold_em_df['季度'] == latest_quarter]
    except KeyError:
        possible_date_cols = [col for col in fund_portfolio_hold_em_df.columns if '季度' in col or 'date' in col.lower() or 'period' in col.lower()]
        if possible_date_cols:
            filtered_df = fund_portfolio_hold_em_df[fund_portfolio_hold_em_df[possible_date_cols[0]] == latest_quarter]
        else:
            return None

    # 获取前十大重仓股
    top_10_stocks = filtered_df.head(10)

    # 获取占净值比例列
    proportion_col = None
    for col in fund_portfolio_hold_em_df.columns:
        if '占净值比例' in str(col) or '净值比例' in str(col) or str(col).endswith('净值比例'):
            proportion_col = col
            break

    # 如果还是找不到，尝试使用列索引（通常是第4列，索引为3）
    if proportion_col is None:
        try:
            proportion_col = fund_portfolio_hold_em_df.columns[3]
        except IndexError:
            return None

    # 计算前十大重仓股的占净值比例之和
    try:
        weight_sum = top_10_stocks[proportion_col].sum()
    except KeyError:
        return None

    return weight_sum


def get_csi_all_share_returns():
    try:
        end_date = datetime.now().strftime('%Y%m%d')
        start_date = (datetime.now() - timedelta(days=3*365)).strftime('%Y%m%d')

        df = ak.stock_zh_index_hist_csindex(symbol='000985', start_date=start_date, end_date=end_date)

        df['日期'] = pd.to_datetime(df['日期'])
        df = df.sort_values('日期', ascending=False).reset_index(drop=True)

        latest_date = df.iloc[0]['日期']
        latest_value = df.iloc[0]['收盘']

        period_mapping = {
            '近1月': 30,
            '近3月': 90,
            '近6月': 180,
            '近1年': 365,
            '近2年': 365 * 2,
            '近3年': 365 * 3
        }

        period_returns = {}

        for period_name, days in period_mapping.items():
            target_date = latest_date - timedelta(days=days)
            target_row = df.iloc[(df['日期'] - target_date).abs().argsort()[:1]]

            if not target_row.empty:
                target_value = target_row.iloc[0]['收盘']
                return_rate = (latest_value - target_value) / target_value * 100
                period_returns[period_name] = round(return_rate, 2)
            else:
                period_returns[period_name] = 0

        start_of_year = datetime(latest_date.year, 1, 1)
        target_row = df.iloc[(df['日期'] - start_of_year).abs().argsort()[:1]]
        if not target_row.empty:
            target_value = target_row.iloc[0]['收盘']
            return_rate = (latest_value - target_value) / target_value * 100
            period_returns['今年来'] = round(return_rate, 2)
        else:
            period_returns['今年来'] = 0

        print(f"中证全指(000985)最新日期：{latest_date.strftime('%Y-%m-%d')}")
        print(f"中证全指各时间段收益率：{period_returns}")

        return period_returns

    except Exception as e:
        print(f"自动获取中证全指数据失败：{e}")
        print("使用默认数据...")
        return {"近3年":25.33,"近2年":49.90,"近1年":33.97,"今年来":5.95,"近6月":8.63,"近3月":-1.35,"近1月":9.57}


def filter_by_excess_return_ratio(row):
    """
    判断超额收益曲线是否"平稳向上"
    
    算法：胜率 + 波动率
    1. 胜率：上升的时间段占比
    2. 波动率：各时间段超额收益的标准差
    3. 综合判断
    
    根据用户例子：
    - A基金：71.65% → 35.62% → 15.68% → 9.09% → 应该通过（都是正的）
    - B基金：3.23% → -5.06% → -5.25% → -4.13% → 应该不通过（持续恶化且变负）
    """
    excess_1y = row["近1年超额"]
    excess_6m = row["近6月超额"]
    excess_3m = row["近3月超额"]
    excess_1m = row["近1月超额"]
    
    # ========== 收集有效数据 ==========
    if pd.isna(excess_1y):
        # 不满一年：只用近6月、近3月、近1月
        excess_list = [excess_6m, excess_3m, excess_1m]
        # 计算年化
        annualized_6m = excess_6m * 2
        annualized_3m = excess_3m * 4
        annualized_1m = excess_1m * 12
        annualized_list = [annualized_6m, annualized_3m, annualized_1m]
    else:
        # 满一年：用全部四个
        excess_list = [excess_1y, excess_6m, excess_3m, excess_1m]
        annualized_1y = excess_1y
        annualized_6m = excess_6m * 2
        annualized_3m = excess_3m * 4
        annualized_1m = excess_1m * 12
        annualized_list = [annualized_1y, annualized_6m, annualized_3m, annualized_1m]
    
    # ========== 指标1：胜率 ==========
    # 计算有多少个时间段是正的
    positive_count = sum(1 for x in excess_list if x > 0)
    win_rate = positive_count / len(excess_list)
    
    # 计算上升的时间段（后面比前面大）
    up_count = 0
    total_compare_pairs = 0
    for i in range(len(excess_list) - 1):
        if excess_list[i + 1] > excess_list[i]:
            up_count += 1
        total_compare_pairs += 1
    improving_rate = up_count / total_compare_pairs if total_compare_pairs > 0 else 0
    
    # ========== 指标2：波动率（标准差） ==========
    # 计算各时间段超额收益的标准差（越小越平稳）
    volatility = np.std(excess_list)
    
    # ========== 指标3：年化收益的稳定性 ==========
    annualized_volatility = np.std(annualized_list)
    
    # ========== 坏的情况（直接不通过） ==========
    # 持续恶化且变负
    if len(excess_list) >= 3:
        if pd.isna(excess_1y):
            # 不满一年的情况
            getting_worse_and_negative = (excess_6m < 0) and (excess_3m < 0) and (excess_1m < 0)
        else:
            getting_worse_and_negative = (excess_1y > 0) and (excess_6m < 0) and (excess_3m < 0)
        
        if getting_worse_and_negative:
            return False
    
    # ========== 好的情况（通过） ==========
    # 条件1：高胜率（大部分时间段都是正的
    high_win_rate = win_rate >= 0.6  # 60%以上是正的
    
    # 条件2：有改善趋势（上升的时间段多）
    has_improvement = improving_rate >= 0.4  # 40%以上的比较是上升的
    
    # 条件3：波动率适中（不是剧烈波动
    low_volatility = volatility < 50  # 标准差小于50%（避免暴涨暴跌）
    
    # 条件4：所有都是正的
    all_positive = all(x > 0 for x in excess_list)
    
    # 条件5：近期表现好
    good_recent = annualized_list[-1] > annualized_list[-2] if len(annualized_list) >= 2 else False
    
    # 组合判断：满足至少两个条件就通过
    score = sum([high_win_rate, has_improvement, low_volatility, all_positive, good_recent])
    
    return score >= 2


def fetch_fund_details(row):
    """单只基金的详细数据获取（用于多线程）"""
    code = row["基金代码"]
    fund_name = row["基金简称"]
    base_name = fund_name[:-1] if (fund_name.endswith('A') or fund_name.endswith('C')) else fund_name
    
    result = {
        "idx": row.name,
        "成立时间": "",
        "最新规模": "",
        "换手率": "",
        "前10大重仓股占比": "",
        "持股行业集中度": "",
        "管理总规模": ""
    }
    
    # 先获取当前基金的基本信息和规模
    try:
        info = get_fund_info(code)
        result["成立时间"] = info['成立时间']
        
        # 先获取当前基金的规模作为基础（支持"亿元"或"亿"格式）
        scale_match = re.search(r'([\d.]+)亿', info['最新规模'])
        total_scale = 0
        if scale_match:
            total_scale = float(scale_match.group(1))
            result["最新规模"] = f"{total_scale:.2f}亿元"
        
        # 再尝试获取A类和C类基金的规模数据并累加
        try:
            code_names = fetch_and_parse_fund_search(base_name)
            if isinstance(code_names, list):
                for code_name in code_names:
                    if isinstance(code_name, dict) and 'code' in code_name and code_name['code'] != code:  # 排除当前基金自己
                        try:
                            info_ac = get_fund_info(code_name['code'])
                            scale_match_ac = re.search(r'([\d.]+)亿', info_ac['最新规模'])
                            if scale_match_ac:
                                total_scale += float(scale_match_ac.group(1))
                                result["最新规模"] = f"{total_scale:.2f}亿元"
                        except:
                            pass
        except Exception as e:
            pass
    except Exception as e:
        pass
    
    # 获取其他详细信息
    try:
        fund_detail = parse_fund_data(code)
        if fund_detail:
            if '换手率' in fund_detail:
                result["换手率"] = fund_detail['换手率']
            if '前10大重仓股占比' in fund_detail:
                result["前10大重仓股占比"] = fund_detail['前10大重仓股占比']
            if '持股行业集中度' in fund_detail:
                result["持股行业集中度"] = fund_detail['持股行业集中度']
            if '管理总规模' in fund_detail:
                result["管理总规模"] = fund_detail['管理总规模']
        
        if not result["前10大重仓股占比"] or pd.isna(result["前10大重仓股占比"]):
            top10_weight = get_top10_stocks_weight_robust(code)
            if top10_weight is not None:
                result["前10大重仓股占比"] = f"{top10_weight:.2f}%"
    except Exception as e:
        pass
    
    return result


def analyze_funds():
    zzqz = get_csi_all_share_returns()

    print("正在获取基金排名数据...")
    fund_open_fund_rank_em_df = ak.fund_open_fund_rank_em(symbol="全部")
    print(f"step1_原始数据：{len(fund_open_fund_rank_em_df)} 只基金")
    fund_open_fund_rank_em_df.to_excel("step1_原始数据.xlsx", index=False)
    
    # 第一步：先做不需要API的快速筛选
    print("正在进行初步筛选...")
    
    # 过滤掉数据为空的（近1年可以为空）
    fund_open_fund_rank_em_df = fund_open_fund_rank_em_df.dropna(subset=['近6月', '近3月', '近1月']).copy()
    print(f"step2_去空值后：{len(fund_open_fund_rank_em_df)} 只基金")
    fund_open_fund_rank_em_df.to_excel("step2_去空值.xlsx", index=False)
    
    # 关键词过滤
    exclude_keywords = ["持有", "A", "通信", "期货", "有色", "黄金", "半导体", "芯片", "云计算", "商品", "创业板",
                        "中证资源", "电信", "物联网", "工程机械", "医药生物", "稀有金属", "科创板",
                        "科创创业", "人工智能", "上海金", "TMT", "指数", "可转债", "债券", "化工", "碳中和", "ESG", "ETF"]
    
    mask = ~fund_open_fund_rank_em_df["基金简称"].str.contains('|'.join(exclude_keywords), na=False)
    fund_open_fund_rank_em_df = fund_open_fund_rank_em_df[mask].copy()
    print(f"step3_关键词过滤后：{len(fund_open_fund_rank_em_df)} 只基金")
    fund_open_fund_rank_em_df.to_excel("step3_关键词过滤.xlsx", index=False)
    
    # 业绩筛选（近1年不强制要求，保留近6月/3月/1月）
    fund_open_fund_rank_em_df = fund_open_fund_rank_em_df[
        (fund_open_fund_rank_em_df["近6月"] >= zzqz["近6月"]) &
        (fund_open_fund_rank_em_df["近3月"] >= zzqz["近3月"]) &
        (fund_open_fund_rank_em_df["近1月"] >= zzqz["近1月"])
    ].copy()
    print(f"step4_业绩筛选后：{len(fund_open_fund_rank_em_df)} 只基金")
    fund_open_fund_rank_em_df.to_excel("step4_业绩筛选.xlsx", index=False)
    
    # 计算超额收益
    fund_open_fund_rank_em_df.loc[:, "近1年超额"] = fund_open_fund_rank_em_df.apply(
        lambda row: row["近1年"] - zzqz["近1年"] if pd.notna(row["近1年"]) else None, 
        axis=1
    )
    fund_open_fund_rank_em_df.loc[:, "近6月超额"] = fund_open_fund_rank_em_df["近6月"] - zzqz["近6月"]
    fund_open_fund_rank_em_df.loc[:, "近3月超额"] = fund_open_fund_rank_em_df["近3月"] - zzqz["近3月"]
    fund_open_fund_rank_em_df.loc[:, "近1月超额"] = fund_open_fund_rank_em_df["近1月"] - zzqz["近1月"]
    
    # 超额收益比例筛选
    # fund_open_fund_rank_em_df = fund_open_fund_rank_em_df[
    #     fund_open_fund_rank_em_df.apply(filter_by_excess_return_ratio, axis=1)
    # ].copy()
    # print(f"step5_超额收益筛选后：{len(fund_open_fund_rank_em_df)} 只基金")
    # fund_open_fund_rank_em_df.to_excel("step5_超额收益筛选.xlsx", index=False)
    
    # 第二步：使用多线程补充详细数据
    fund_open_fund_rank_em_df.loc[:, "成立时间"] = ""
    fund_open_fund_rank_em_df.loc[:, "最新规模"] = ""
    fund_open_fund_rank_em_df.loc[:, "换手率"] = ""
    fund_open_fund_rank_em_df.loc[:, "前10大重仓股占比"] = ""
    fund_open_fund_rank_em_df.loc[:, "持股行业集中度"] = ""
    
    print(f"正在使用多线程补充基金详细数据（线程数：20）...")
    
    # 使用ThreadPoolExecutor并行获取数据
    with ThreadPoolExecutor(max_workers=20) as executor:
        # 提交所有任务
        futures = {executor.submit(fetch_fund_details, row): idx for idx, row in fund_open_fund_rank_em_df.iterrows()}
        
        # 收集结果
        for future in tqdm(as_completed(futures), total=len(futures)):
            try:
                result = future.result()
                idx = result["idx"]
                fund_open_fund_rank_em_df.at[idx, "成立时间"] = result["成立时间"]
                fund_open_fund_rank_em_df.at[idx, "最新规模"] = result["最新规模"]
                fund_open_fund_rank_em_df.at[idx, "换手率"] = result["换手率"]
                fund_open_fund_rank_em_df.at[idx, "前10大重仓股占比"] = result["前10大重仓股占比"]
                fund_open_fund_rank_em_df.at[idx, "持股行业集中度"] = result["持股行业集中度"]
                fund_open_fund_rank_em_df.at[idx, "管理总规模"] = result["管理总规模"]
            except Exception as e:
                pass
    
    # 保存补充完详细数据的结果
    fund_open_fund_rank_em_df.to_excel("step6_补充数据完成.xlsx", index=False)
    
    # 第三步：基于补充的数据做最终筛选
    print("正在进行最终筛选...")
    
    # 打印规模数据统计
    scale_data = fund_open_fund_rank_em_df["最新规模"].values
    empty_count = sum(1 for s in scale_data if s == "" or pd.isna(s))
    print(f"规模数据情况：总共 {len(scale_data)} 只，空数据 {empty_count} 只，有数据 {len(scale_data) - empty_count} 只")
    if len(scale_data) - empty_count > 0:
        sample_scales = [s for s in scale_data if s != "" and not pd.isna(s)][:10]
        print(f"规模数据示例：{sample_scales}")
    
    # 换手率筛选
    def filter_by_turnover(turnover_str):
        if pd.isna(turnover_str) or turnover_str == "" or turnover_str == "获取失败":
            return True
        if "%" in turnover_str:
            turnover_value = float(turnover_str.replace("%", ""))
            return turnover_value >= 500
        return False
    
    fund_open_fund_rank_em_df = fund_open_fund_rank_em_df[
        fund_open_fund_rank_em_df["换手率"].apply(filter_by_turnover)
    ].copy()
    print(f"step7_换手率筛选后：{len(fund_open_fund_rank_em_df)} 只基金")
    fund_open_fund_rank_em_df.to_excel("step7_换手率筛选.xlsx", index=False)
    
    # 规模筛选
    debug_scale_results = []
    def filter_by_scale(scale_str):
        try:
            if pd.isna(scale_str) or scale_str == "" or scale_str == "获取失败":
                debug_scale_results.append((scale_str, "空数据"))
                return False  # 空数据直接排除
            
            # 清理一下字符串
            scale_str_clean = str(scale_str).strip()
            
            # 统一转换为亿元单位
            if "亿" in scale_str_clean:
                # 提取数字部分
                num_match = re.search(r'([\d.]+)', scale_str_clean)
                if num_match:
                    scale_value = float(num_match.group(1))
                else:
                    debug_scale_results.append((scale_str, "无法提取数字"))
                    return False
            elif "万" in scale_str_clean:
                # 万元转亿元：除以10000
                num_match = re.search(r'([\d.]+)', scale_str_clean)
                if num_match:
                    scale_value = float(num_match.group(1)) / 10000
                else:
                    debug_scale_results.append((scale_str, "无法提取数字"))
                    return False
            else:
                # 假设无单位的数据是亿元
                num_match = re.search(r'([\d.]+)', scale_str_clean)
                if num_match:
                    scale_value = float(num_match.group(1))
                else:
                    debug_scale_results.append((scale_str, "无法提取数字"))
                    return False
            
            # 判断是否在范围内
            in_range = 0.2 <= scale_value <= 10
            debug_scale_results.append((scale_str, scale_value, in_range))
            return in_range
        except Exception as e:
            debug_scale_results.append((scale_str, f"异常：{e}"))
            return False  # 解析失败排除
    
    fund_open_fund_rank_em_df = fund_open_fund_rank_em_df[
        fund_open_fund_rank_em_df["最新规模"].apply(filter_by_scale)
    ].copy()
    print(f"step8_规模筛选后：{len(fund_open_fund_rank_em_df)} 只基金")
    fund_open_fund_rank_em_df.to_excel("step8_规模筛选.xlsx", index=False)
    
    # 打印调试信息
    print("\n=== 规模筛选调试信息 ===")
    for i, res in enumerate(debug_scale_results[:30]):  # 只打印前30条
        print(f"{i+1}. {res}")
    if len(debug_scale_results) > 30:
        print(f"... 还有 {len(debug_scale_results)-30} 条")
    
    # 前10大重仓股占比筛选（<40%）
    print("\n=== 正在进行前10大重仓股筛选 ===")
    debug_holdings_results = []
    
    def filter_by_holdings(holdings_str):
        try:
            if pd.isna(holdings_str) or holdings_str == "" or holdings_str == "获取失败":
                debug_holdings_results.append((holdings_str, "空数据"))
                return True  # 空数据保留
            
            # 提取数字
            num_match = re.search(r'([\d.]+)', str(holdings_str))
            if num_match:
                holdings_value = float(num_match.group(1))
                in_range = holdings_value < 40
                debug_holdings_results.append((holdings_str, holdings_value, in_range))
                return in_range
            else:
                debug_holdings_results.append((holdings_str, "无法提取数字"))
                return True  # 无法解析的保留
        except Exception as e:
            debug_holdings_results.append((holdings_str, f"异常：{e}"))
            return True  # 异常保留
    
    fund_open_fund_rank_em_df = fund_open_fund_rank_em_df[
        fund_open_fund_rank_em_df["前10大重仓股占比"].apply(filter_by_holdings)
    ].copy()
    print(f"step9_前10大重仓股筛选后：{len(fund_open_fund_rank_em_df)} 只基金")
    fund_open_fund_rank_em_df.to_excel("step9_前10大重仓股筛选.xlsx", index=False)
    
    # 打印调试信息
    print("\n=== 前10大重仓股筛选调试信息 ===")
    for i, res in enumerate(debug_holdings_results[:30]):
        print(f"{i+1}. {res}")
    if len(debug_holdings_results) > 30:
        print(f"... 还有 {len(debug_holdings_results)-30} 条")
    
    # 持股行业集中度筛选（<40%）
    print("\n=== 正在进行持股行业集中度筛选 ===")
    debug_concentration_results = []
    
    def filter_by_concentration(concentration_str):
        try:
            if pd.isna(concentration_str) or concentration_str == "" or concentration_str == "获取失败":
                debug_concentration_results.append((concentration_str, "空数据"))
                return True  # 空数据保留
            
            # 提取数字
            num_match = re.search(r'([\d.]+)', str(concentration_str))
            if num_match:
                concentration_value = float(num_match.group(1))
                in_range = concentration_value < 40
                debug_concentration_results.append((concentration_str, concentration_value, in_range))
                return in_range
            else:
                debug_concentration_results.append((concentration_str, "无法提取数字"))
                return True  # 无法解析的保留
        except Exception as e:
            debug_concentration_results.append((concentration_str, f"异常：{e}"))
            return True  # 异常保留
    
    fund_open_fund_rank_em_df = fund_open_fund_rank_em_df[
        fund_open_fund_rank_em_df["持股行业集中度"].apply(filter_by_concentration)
    ].copy()
    print(f"step10_持股行业集中度筛选后：{len(fund_open_fund_rank_em_df)} 只基金")
    fund_open_fund_rank_em_df.to_excel("step10_持股行业集中度筛选.xlsx", index=False)
    
    # 打印调试信息
    print("\n=== 持股行业集中度筛选调试信息 ===")
    for i, res in enumerate(debug_concentration_results[:30]):
        print(f"{i+1}. {res}")
    if len(debug_concentration_results) > 30:
        print(f"... 还有 {len(debug_concentration_results)-30} 条")
    
    print(f"\n最终筛选后剩余 {len(fund_open_fund_rank_em_df)} 只基金")
    
    # 删除不需要的列
    columns_to_drop = ["序号", "单位净值", "累计净值", "日增长率", "自定义", "手续费"]
    # 只删除实际存在的列
    existing_columns = [col for col in columns_to_drop if col in fund_open_fund_rank_em_df.columns]
    df_export = fund_open_fund_rank_em_df.drop(columns=existing_columns, errors='ignore')
    
    # 按近6月超额倒序排序
    if "近6月超额" in df_export.columns:
        df_export = df_export.sort_values(by="近6月超额", ascending=False)
    
    # 使用openpyxl优化Excel格式
    
    # 保存到两个文件
    output_files = ["step11_最终结果.xlsx", "fund_open_fund_rank_em.xlsx"]
    
    for file in output_files:
        df_export.to_excel(file, index=False)
        
        # 打开文件并优化格式
        wb = load_workbook(file)
        ws = wb.active
        
        # 设置所有单元格居中对齐
        alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment
        
        # 自适应列宽（紧凑版）
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # 获取列字母
            column_name = col[0].value
            
            for cell in col:
                try:
                    # 计算字符长度（中文字符按1.5个宽度算，更贴近Excel实际显示）
                    cell_value = str(cell.value)
                    length = 0
                    for char in cell_value:
                        if '\u4e00' <= char <= '\u9fff':  # 中文字符
                            length += 1.5
                        else:
                            length += 1
                    if length > max_length:
                        max_length = length
                except:
                    pass
            
            # 计算列宽（更紧凑）
            adjusted_width = max_length + 1  # 加1个字符的padding就够
            
            # 特殊列适当加宽，但不要太大
            if column_name == "基金简称":
                adjusted_width = max(adjusted_width, 12)  # 基金简称至少12
            elif column_name == "基金代码":
                adjusted_width = max(adjusted_width, 10)  # 基金代码至少10
            elif "超额" in str(column_name) or "收益率" in str(column_name):
                adjusted_width = max(adjusted_width, 10)
            
            # 最大宽度限制（缩紧到40）
            adjusted_width = min(adjusted_width, 40)
            
            ws.column_dimensions[column].width = adjusted_width
        
        wb.save(file)
    
    print("结果已保存到 fund_open_fund_rank_em.xlsx（格式已优化）")


if __name__ == "__main__":
    analyze_funds()
